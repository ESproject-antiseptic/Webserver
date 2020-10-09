from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse,HttpResponse
from django.shortcuts import render,redirect
from .models import *
from main.models import*
from . import views
from django.contrib import messages
import simplejson
from django.utils.decorators import method_decorator
from django.contrib.auth.hashers import make_password,check_password

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

# Create your views here.
def home(request):
    if request.session.get('room'):
        del(request.session['room'])
    return render(request, 'home/home.html')

from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def make_room(request): #방만들기
    info = {}
    if request.method == "GET":
        room_name = request.GET.get('room_name')
        if room_name:
            try: #룸이름이 중복되는 경우
                exist = Room.objects.get(room_name =room_name)
                overlap = "fail"
                info['overlap'] = overlap
            except Room.DoesNotExist:
                info['roomcheck']=room_name
                overlap="pass"
                info['orverlap']=overlap
            return JsonResponse(info)
        else:
            print('GET & else')
            return render(request, 'home/make_room.html')

    elif request.method == "POST":
        user_email = request.session.get('user')
        admin = User.objects.get(email=user_email)
        room_name = request.POST.get('room_name')
        room_ps=request.POST.get('room_ps')
        try:
            room_member =request.FILES['room_member'] #명단
        except :
            print('none')
            room_member= None
        check_values=request.POST.getlist('chBox') #2.특정프로그램차단 3.WEB키보드 사용금지 4. APP 움직임 감지 5.마스크인식
        print(check_values)
        room_func = [] #기능들 번호 리스트
        member_list = []
        if room_member:#chBox1이면 'room_member'(명단) =명단 있어서 얼굴인식기능 선택
            print('chbox1 존재')
            room_func.append('1')
            print(room_func)
            print('hhh')
            #명단에서학번만 추출
            fs = FileSystemStorage()
            filename = fs.save(room_member.name, room_member)
            uploaded_file_url=fs.url(filename)
            print(uploaded_file_url)

            print('hhhhh')
            member = load_workbook("media/" + room_member.name)
            sheet = member['Sheet1']
            col=sheet['A']#A열 학번데이터 가져오기

            for cell in col:
                member_list.append(cell.value)
            print(member_list)

        room_func.extend(check_values)
        print(room_func)

        if not (room_name and room_ps and room_func):
            messages.add_message(request, messages.INFO, '모든 값을 입력해야 합니다.!')  # 첫번째, 초기지원
            return render(request, 'home/make_room.html')
        else:
            info['admin']=admin
            info['room_name']=room_name
            info['room_ps']=room_ps
            room = Room(admin=admin, room_name=room_name, room_ps=make_password(room_ps), room_member=room_member,room_func=room_func , room_member_list=member_list)
            room.save()
            messages.add_message(request, messages.INFO, '방만들기가 성공하였습니다.')


            return render(request, 'home/make_room_success.html',info) #성공

#방들어가기
@csrf_exempt#사용안함
def enter_room(request):
    if request.method == "GET":
        if request.session.get('room'):
            del (request.session['room'])
        return render(request,'home/enter_room.html')

    elif request.method == "POST":
        room_name = request.POST.get('room_name')
        room_ps = request.POST.get('room_ps')

        try:
            room = Room.objects.get(room_name=room_name) #입력받은 room_name에 해당하는 방이 있는지 DB에서 검색
            if check_password(room_ps, room.room_ps): #방이 존재한다면 해당 방의 room_ps와 입력받은 비번 검사
                request.session['room']:room #room.room_name이라고해도 무방
                messages.add_message(request, messages.INFO, '방입장 성공')
                print('hihihi')
                print(room.room_func)
                #room 기능 속에 1번기능이 있다면
                if '1' in room.room_func: #room_func은 리스트형식
                    return redirect("/home/enter_room_recognition") #얼굴인식페이지로
                #1번기능 없다면(얼굴인식 필요없음!!)
                else:
                    return redirect('room',room_name) #url이랑 여기 맞추는거 다시고쳐야함!

            else:
                messages.add_message(request,messages.INFO,'비밀번호가 틀렸습니다')

        except Room.DoesNotExist: #room_name이없을때
            messages.add_message(request.messages.INFO, '잘못된 방 코드입니다.')



        return render(request, 'home/enter_room.html')

#ajax로 방들어가기
def enter_room2(request):
    if request.method == "GET":
        if request.session.get('room'):
            del (request.session['room'])
        #ajax였으면
        if (request.GET.get('room_name')):
            room_name = request.GET.get('room_name')
            room_ps= request.GET.get('room_ps')
            info = {}
            try:
                room = Room.objects.get(room_name=room_name)  # 입력받은 room_name에 해당하는 방이 있는지 DB에서 검색
                if check_password(room_ps, room.room_ps):  # 방이 존재한다면 해당 방의 room_ps와 입력받은 비번 검사
                    request.session['room']=room_name  # room.room_name이라고해도 무방
                    print('room')
                    info['room_name']=room_name
                    messages.add_message(request, messages.INFO, '방입장 성공')
                    # room 기능 속에 1번기능이 있다면
                    if '1' in room.room_func:  # room_func은 리스트형식
                        info['confirm'] = "need"
                        return JsonResponse(info)# 얼굴인식페이지로
                    # 1번기능 없다면(얼굴인식 필요없음!!)
                    else:
                        info['confirm'] = "goroom"
                        return JsonResponse(info)  #
                else:
                    info['confirm'] = "비밀번호가 틀렸습니다"
                    return JsonResponse(info)
            except Room.DoesNotExist:  # room_name이없을때
                info['confirm'] = "잘못된 방 코드입니다."
                return JsonResponse(info)
        else:
            return render(request,'home/enter_room2.html')

    elif request.method == "POST":
        return render(request, 'home/enter_room2.html')



from luxand import luxand
#방 입장시 얼굴인식 페이지
@csrf_exempt
def enter_room_recognition(request):
    if request.method == "GET":
        return render(request, 'home/enter_room_recognition.html')

    elif request.method == "POST":
        info={}
        member_number = request.POST.get('member_number')
        member_name = request.POST.get('member_name')
        member_image = request.FILES['photo']

        fs = FileSystemStorage()
        member_image_path = fs.url(member_image.name)
        print(member_image_path)
        print('hyewonhihihi')

        #해당방의 DB속 명단Excel파일 조회
        room_name=request.session.get('room')
        room = Room.objects.get(room_name=room_name)
        member_file=room.room_member #명단

        #DB의 member_list로 회원번호 확인 및 index 추출
        member_list=room.room_member_list #회원번호만 적힌 리스트
        member_list= member_list[1:-1].split(', ')
        print(member_list)
        if (member_number in member_list):
            member_index=member_list.index(member_number) +1 #index=0은'회원번호(수험번호/학번)'이므로 index로 추출된 수 +1로 쓰면됨!
            print('okokokok')
            member = load_workbook("media/" + member_file.name)
            sheet = member['Sheet1']
            #B열은 회원이름, C열은 사진
            member_file_name = sheet['B'+str(member_index)].value #이러케표현도가능 sheet.cell(2,1).value
            #이름이 틀리면
            if member_file_name != member_name:
                info['result'] = "NO_NAME"
                print('no_name')
            #이름이 같다면 사진 비교할것
            else:
                image_loader = SheetImageLoader(sheet)
                image = image_loader.get('C'+str(member_index))
                member_file_image_path = ("media/c"+str(member_number)+".jpg")
                image.save(member_file_image_path)
                a="/Users/Owner/PycharmProjects/Webserver/Webserver/antiseptic19/"+str(member_file_image_path)
                b="/Users/Owner/PycharmProjects/Webserver/Webserver/antiseptic19"+str(member_image_path)
                #회원이름 비교 및 사진 비교
                client = luxand("12a42a8efedf4e24b84730ce440e5429")
                member_file_image = client.add_person(str(member_file_name), photos=[a])
                result = client.verify(member_file_image, photo=b)
                print(result)
                #인증 성공
                if result['status']=='success':
                    info['result']="OK"
                    print("ok")
                else:
                    info['result']="NO_IMAGE_MATCH"
        else:
            #명단 속 존재하지 않는 회원번호 (입장불가!)
            info['result']="NO_MEMBER"
            print('no_member')
            #해당페이지에서 팝업으로 입장불가표시주기
        return render(request, 'home/enter_room_recognition.html',info)


        # if member_number in room.room_member_list:
        #
        # image_loader = SheetImageLoader(sheet)
        # image = image_loader.get('C2')
        # image.show()
        # image.save('media/c2.jpg')
        # if  #얼굴인식 완료되면
        return render(request, 'home/enter_room_recognition.html')

#사용안함
#ajax로하는거 (url에 특정 방이름들어가면 -->data 주고받는게 안되더라.. ajax도 불가, form POST or GET도 불가!!
def recognition(request,room_name):
    if request.method == "GET":
        print('hihi')
        return render(request, 'home/recognition.html')

    elif request.method == "POST":
        info ={}
        member_number = request.POST.get('member_number')
        member_name = request.POST.get('member_name')

        #해당방의 DB속 명단Excel파일 조회
        room_name=request.session.get('room')
        room = Room.objects.get(room_name=room_name)
        member_file=room.room_member #명단
        member_list=room.room_member_list
        print(member_file)
        print(member_list)
        print(member_list.type)
        if (member_number in member_list):
            print('okokokok')
        # if member_number in room.room_member_list:
        #
        # member = load_workbook("media/"+member_file.name)
        # sheet = member['Sheet1']
        # a=sheet['A1'].value
        # aa=sheet.cell(2,1).value
        # print(a)
        # print(aa)
        # print('hello')
        # print(member.rows.values)
        #
        # image_loader = SheetImageLoader(sheet)
        # image = image_loader.get('C2')
        # image.show()
        # image.save('media/c2.jpg')
        # if  #얼굴인식 완료되면
        info['result']='ok'
        return JsonResponse(info)




#특정방
def room(request,room_name):
    print('hihi')
    return render(request, 'home/room.html')


def myroom(request):
    return render(request, 'home/myroom.html')

def myroom1(request):
    user_email=request.session.get('user')
    admin = User.objects.get(email=user_email)
    room = Room.objects.filter(admin=admin)
    rooms= room.order_by('-id')
    return render(request, 'home/myroom1.html', {"rooms":rooms})





#앱 이름 비번 명단 모두 제출시
@method_decorator(csrf_exempt,name='dispatch')
def app_makeroom(request):
    if request.method == "GET":
        return render(request, 'home/make_room.html')
    if request.method =="POST":
        #res_data = {}
        myfile = request.FILES['files']
        myemail = request.POST.get('admin') #아이디(이멜)
        myadmin = User.objects.get(email=myemail)
        myname = request.POST.get('name')
        mypass = request.POST.get('pass')
        mycheckbox = request.POST.get('checkbox')
        print(myname)
        #fs = FileSystemStorage()
        #res_data['file_url'] = fs.url(file.name)
        print(myfile)
        myuser = Room(room_name=myname,room_ps=mypass,admin=myadmin,room_func=mycheckbox)
        myuser.room_member = myfile
        myuser.save()
        return HttpResponse(simplejson.dumps({"file":myfile.name}))
#앱 이름 비번만 제출시
@method_decorator(csrf_exempt,name='dispatch')
def app_makemyroom(request):
    if request.method == "GET":
        return render(request, 'home/make_room.html')
    if request.method =="POST":
        #res_data = {}
        myemail = request.POST.get('admin') #아이디(이멜)
        myadmin = User.objects.get(email=myemail)
        roomname = request.POST.get('roomname')
        password = request.POST.get('password')
        mycheckbox = request.POST.get('checkbox')
        myuser = Room(room_name=roomname,room_ps=make_password(password),admin=myadmin,room_func=mycheckbox)
        myuser.save()
        print(myuser)
        return HttpResponse(simplejson.dumps({"roomname":roomname,"password":password}))


#앱 중복체크
@method_decorator(csrf_exempt,name='dispatch')
def app_roomnumber(request):
    if request.method =="POST":
        roomname = request.POST.get("roomname")
        print(roomname)
        if Room.objects.filter(room_name=roomname).exists():
            return HttpResponse(simplejson.dumps({"roomname":"exist"}))
        else:
            return HttpResponse(simplejson.dumps({"roomname":roomname}))
